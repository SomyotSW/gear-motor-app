"""app.py
AC Gear Motor Quote PDF Generator
- Reads price from: AC Gear Motor Price List 2026 (sheet1)
- Fills template: QMO26-SAS.xlsx (sheet: Sales Quote  (2))
- Converts filled XLSX -> PDF using LibreOffice (soffice)
- Returns PDF for download
- Saves PDF to output_pdfs/ and auto-deletes files older than 7 days
- (Optional) Emails the PDF to customer via SMTP if SMTP env vars are configured

Minimal-impact design:
- Only backend responsibilities; frontend can keep downloading PDF blob.
"""

from __future__ import annotations

import os
import re
import time
import shutil
import smtplib
import tempfile
import subprocess
from datetime import datetime   
from pathlib import Path
from email.message import EmailMessage
try:
    import fcntl  # Render/Linux มี
except ImportError:
    fcntl = None  # Windows local dev อาจไม่มี
from flask import Flask, request, send_file, jsonify, abort
from openpyxl import load_workbook

from flask_cors import CORS
import requests as http_requests
import hashlib
import hmac
import base64

app = Flask(__name__)

CORS(
    app,
    resources={r"/api/*": {"origins": [
        "https://sas-gear-motor-app.vercel.app",
        "https://sas-gear-motor-q9s9l76vq-somyot442s-projects.vercel.app",
    ]}},
    methods=["GET", "POST", "OPTIONS"],
    allow_headers=["Content-Type", "Authorization"],
)

# =========================
# CORS (dev + vercel)
# - Reflect Origin if it's in allowlist
# - Always allow OPTIONS preflight for /api/*
# =========================
ALLOWED_ORIGINS = {
    "http://localhost:3000",
    "http://127.0.0.1:3000",
    "https://sas-gear-motor-app.vercel.app",
}

from flask import request

@app.after_request
def add_cors_headers(response):
    if request.path.startswith("/line/"):
        return response
    origin = request.headers.get("Origin")

    # ✅ ADD: allow vercel preview domains too
    is_vercel_preview = (
        origin
        and origin.endswith(".vercel.app")
        and ("sas-gear-motor" in origin)
    )

    if origin in ALLOWED_ORIGINS or is_vercel_preview:
        response.headers["Access-Control-Allow-Origin"] = origin
        response.headers["Vary"] = "Origin"

    response.headers["Access-Control-Allow-Methods"] = "GET,POST,OPTIONS"
    # ✅ ADD: allow common headers used by browsers
    response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
    response.headers["Access-Control-Expose-Headers"] = "Content-Disposition"
    return response

@app.before_request
def handle_preflight():
    if request.path.startswith("/line/"):
        return None
    if request.method == "OPTIONS" and request.path.startswith("/api/"):
        return ("", 204)
    
@app.route("/api/ac-quote", methods=["OPTIONS"])
def ac_quote_preflight():
    return ("", 204)

# =========================
# Paths
# =========================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")

PRICE_FILE = os.path.join(DATA_DIR, "AC Gear Motor Price List 2026 14-02-26.xlsx")
TEMPLATE_FILE = os.path.join(DATA_DIR, "QMO26-SAS.xlsx")

# BLDC price file (ไฟล์แยกต่างหากจาก AC)
BLDC_PRICE_FILE = os.path.join(DATA_DIR, "High BLDC Gear Motor Price List 2026 14-02-26.xlsx")

PRICE_SHEET_NAME = "Sheet1"              # หากชื่อชีทจริงไม่ใช่ sheet1 ให้แก้ตรงนี้
TEMPLATE_SHEET_NAME = "Sales Quote  (2)" # ชื่อชีทใน QMO26-SAS.xlsx
# เพิ่มใหม่ — ข้อมูล Sale Person
SALE_PERSONS = {
    'CA':  {
        'abbr': 'CA',
        'name': 'Mr. Chottanin A. (CA)',
        'position': 'TRANSMISSION PRODUCT MANAGER : 081-921-6225'
    },
    'AP': {
        'abbr': 'AP',
        'name': 'Ms.Apichaya P. (AP)',
        'position': 'Sale Supervisor 098-3697494'
    },
    'MY': {
        'abbr': 'MY',
        'name': 'Ms.Matavee Y. (MY)',
        'position': 'Sale Supervisor 092-2715371'
    },
    'TWS': {
        'abbr': 'TWS',
        'name': 'Ms.Thitikan W. (TWS)',
        'position': 'Sale Exclusive 080-4632394'
    },
    'PW':  {
        'abbr': 'PW',
        'name': 'Mr.Parada W.(PW)',
        'position': 'Sale Engineer 088-9404948'
    },
    'SI':  {
        'abbr': 'SI',
        'name': 'Ms.Suphak I.(SI)',
        'position': 'Sale Exclusive 096-0787776'
    },
    'NM':  {
        'abbr': 'NM',
        'name': 'Mr.Naphaphat M.(NM)',
        'position': 'Sale Exclusive 065-7176332'
    },
    'SK':  {
        'abbr': 'SK',
        'name': 'Mr.Sanya K.(SK)',
        'position': 'Sale Supervisor 086-9819616'
    },
    'PL':  {
        'abbr': 'PL',
        'name': 'Mr.Pongsakorn L.(PL)',
        'position': 'Sale Engineer 063-2159056'
    },
    'TL':  {
        'abbr': 'TL',
        'name': 'Ms.Tanawee L.(TL)',
        'position': 'Sale Supervisor 092-2715372'
    },
    'NR':  {
        'abbr': 'NR',
        'name': 'Ms.Nantida R.(NR)',
        'position': 'Sale Exclusive 098-2711425'
    },
}

# =========================
# Internal notify recipients (always CC'd when quote PDF is generated)
# =========================
INTERNAL_NOTIFY_EMAILS = [
    "Chottanin@synergy-as.com",
    "sas04@synergy-as.com",
]

# =========================
# Output storage (7 days retention)
# =========================
OUTPUT_DIR = Path(os.path.join(BASE_DIR, "output_pdfs"))
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
RETENTION_SECONDS = 7 * 24 * 60 * 60
# ===== ADD: running quotation number (QMO26-SAS-001...) =====
SEQ_FILE = os.environ.get("QMO_SEQ_FILE", str(OUTPUT_DIR / "qmo_seq.txt"))

def next_qmo_no() -> int:
    os.makedirs(os.path.dirname(SEQ_FILE), exist_ok=True)

    with open(SEQ_FILE, "a+", encoding="utf-8") as f:
        if fcntl:
            fcntl.flock(f, fcntl.LOCK_EX)

        f.seek(0)
        raw = f.read().strip()
        last_no = int(raw) if raw.isdigit() else 0
        new_no = last_no + 1

        f.seek(0)
        f.truncate()
        f.write(str(new_no))
        f.flush()

        if fcntl:
            fcntl.flock(f, fcntl.LOCK_UN)

    return new_no

def cleanup_old_pdfs() -> None:
    now = time.time()
    for p in OUTPUT_DIR.glob("*.pdf"):
        try:
            if p.is_file() and (now - p.stat().st_mtime) > RETENTION_SECONDS:
                p.unlink(missing_ok=True)
        except Exception:
            # don't break requests because cleanup failed
            pass

# =========================
# Price map helpers
# =========================

def norm_code(x: str) -> str:
    s = str(x or "")
    s = s.replace("\u00A0", "")   # NBSP
    s = s.replace("\u2011", "-")  # non-breaking hyphen → hyphen
    s = s.replace("\u2013", "-")  # en-dash → hyphen
    s = s.replace("\u2014", "-")  # em-dash → hyphen
    s = s.strip().upper()
    s = re.sub(r"\s+", "", s)
    return s

def build_price_map() -> dict[str, dict[str, object]]:
    if not os.path.exists(PRICE_FILE):
        raise FileNotFoundError(f"Price file not found: {PRICE_FILE}")

    wb = load_workbook(PRICE_FILE, data_only=True)
    ws = wb[PRICE_SHEET_NAME] if PRICE_SHEET_NAME in wb.sheetnames else wb.active

    out: dict[str, dict[str, object]] = {}
    max_row = ws.max_row or 0
    for r in range(1, max_row + 1):
        code = ws.cell(r, 1).value  # Col A
        if not code:
            continue
        raw = str(code).strip()
        key = norm_code(raw)

        desc = ws.cell(r, 2).value  # Col B
        price = ws.cell(r, 4).value # Col D

        try:
            price_val = float(price) if price not in (None, "") else 0.0
        except Exception:
            price_val = 0.0

        out[key] = {
            "raw": raw,
            "desc": str(desc).strip() if desc is not None else "",
            "price": price_val,
        }

    return out

try:
    PRICE_MAP = build_price_map()
except Exception as e:
    PRICE_MAP = {}
    print(f"[WARN] Cannot load PRICE_MAP: {e}")

@app.get("/api/ac-price-reload")
def reload_price():
    global PRICE_MAP
    try:
        PRICE_MAP = build_price_map()
        return jsonify({"ok": True, "count": len(PRICE_MAP)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

# =========================
# XLSX -> PDF via LibreOffice
# =========================

def find_soffice() -> str:
    # 1) explicit env override
    env_path = os.environ.get("SOFFICE_PATH", "").strip()
    if env_path and os.path.exists(env_path):
        return env_path

    # 2) PATH lookup (Linux / some Windows setups)
    p = shutil.which("soffice")
    if p:
        return p

    # 3) common Windows default install paths
    candidates = [
        r"C:\\Program Files\\LibreOffice\\program\\soffice.exe",
        r"C:\\Program Files (x86)\\LibreOffice\\program\\soffice.exe",
    ]
    for c in candidates:
        if os.path.exists(c):
            return c

    raise FileNotFoundError(
        "LibreOffice soffice not found. Install LibreOffice or set SOFFICE_PATH env var."
    )

def xlsx_to_pdf(xlsx_path: str, out_dir: str) -> str:
    soffice = find_soffice()

    subprocess.check_call([
        soffice,
        "--headless",
        "--nologo",
        "--nolockcheck",
        "--convert-to", "pdf",
        "--outdir", out_dir,
        xlsx_path,
    ])

    base = os.path.splitext(os.path.basename(xlsx_path))[0]
    pdf_path = os.path.join(out_dir, base + ".pdf")
    if not os.path.exists(pdf_path):
        raise FileNotFoundError("PDF convert failed: output pdf not created.")
    return pdf_path

# =========================
# Optional email sending (SMTP)
# =========================

def smtp_is_configured() -> bool:
    return all(os.environ.get(k) for k in ["SMTP_HOST", "SMTP_USER", "SMTP_PASS"]) 

def send_email_with_pdf(to_email: str, subject: str, body: str, pdf_path: str) -> None:
    SMTP_HOST = os.environ.get("SMTP_HOST", "")
    SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
    SMTP_USER = os.environ.get("SMTP_USER", "")
    SMTP_PASS = os.environ.get("SMTP_PASS", "")
    FROM_EMAIL = os.environ.get("FROM_EMAIL", SMTP_USER)

    if not (SMTP_HOST and SMTP_USER and SMTP_PASS and FROM_EMAIL):
        raise RuntimeError("SMTP env not set (SMTP_HOST/SMTP_USER/SMTP_PASS/FROM_EMAIL)")

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = FROM_EMAIL
    msg["To"] = to_email
    msg.set_content(body)

    with open(pdf_path, "rb") as f:
        pdf_bytes = f.read()

    msg.add_attachment(
        pdf_bytes,
        maintype="application",
        subtype="pdf",
        filename=os.path.basename(pdf_path),
    )

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
        s.starttls()
        s.login(SMTP_USER, SMTP_PASS)
        s.send_message(msg)

# =========================
# Download endpoint (served from output_pdfs)
# =========================

@app.get("/api/download/<filename>")
def download_pdf(filename: str):
    # [BUG FIX #6] ป้องกัน path traversal attack
    p = (OUTPUT_DIR / filename).resolve()
    if not str(p).startswith(str(OUTPUT_DIR.resolve())):
        abort(400)
    if not p.exists():
        abort(404)
    return send_file(str(p), mimetype="application/pdf", as_attachment=True, download_name=filename)

# =========================
# Main quote endpoint
# =========================

@app.post("/api/ac-quote")
def ac_quote():
    try:
        app.logger.info("RUNNING FILE = %s", __file__)
        APP_VERSION = "2026-03-03-fix-sheetprops-v1"
        app.logger.info("APP_VERSION = %s", APP_VERSION)
        payload = request.get_json(silent=True) or {}

        motor_code = str(payload.get("motorCode", "")).strip()
        gear_code = str(payload.get("gearCode", "")).strip()

        try:
            qty_motor = int(payload.get("qtyMotor", 1) or 1)
        except Exception:
            qty_motor = 1

        try:
            qty_gear = int(payload.get("qtyGear", 1) or 1)
        except Exception:
            qty_gear = 1
        
        ctrl_model = str(payload.get("controllerModel", "") or "").strip()
        try:
            qty_ctrl = int(payload.get("qtyCtrl", 0) or 0)
        except Exception:
            qty_ctrl = 0

        if not motor_code or not gear_code:
            return "Invalid motorCode/gearCode", 400

        if not os.path.exists(TEMPLATE_FILE):
            return f"Template file not found: {TEMPLATE_FILE}", 500

        if not PRICE_MAP:
            return "PRICE_MAP is empty. Check PRICE_FILE path or reload /api/ac-price-reload", 500
        
        ctrl = None
        if ctrl_model and qty_ctrl > 0:
            ckey = norm_code(ctrl_model)
            if ckey not in PRICE_MAP:
                return f"Controller code not found in price list: {ctrl_model}", 404
            ctrl = PRICE_MAP[ckey]

        mkey = norm_code(motor_code)
        gkey = norm_code(gear_code)

        if mkey not in PRICE_MAP:
            return f"Motor code not found in price list: {motor_code}", 404
        if gkey not in PRICE_MAP:
            return f"Gear code not found in price list: {gear_code}", 404

        motor = PRICE_MAP[mkey]
        gear = PRICE_MAP[gkey]

        wb = load_workbook(TEMPLATE_FILE)
        if TEMPLATE_SHEET_NAME not in wb.sheetnames:
            return f"Template sheet not found: {TEMPLATE_SHEET_NAME}", 500

        ws = wb[TEMPLATE_SHEET_NAME]
        if ws is None:
            return jsonify({"ok": False, "error": "Worksheet is None (failed to load template sheet)"}), 500

        # ===== ADD: customer fields -> template cells =====
        cust = payload.get("customer", {}) or {}
        cust_name = str(cust.get("name", "")).strip()
        cust_company = str(cust.get("company", "")).strip()
        cust_phone = str(cust.get("phone", "")).strip()
        cust_email = str(cust.get("email", "")).strip()

        # ===== salePerson fields =====
        sale_person_abbr = str(payload.get("salePerson", "CA")).strip() or "CA"
        sp = SALE_PERSONS.get(sale_person_abbr, SALE_PERSONS["CA"])

        # B8: "คุณ : ..."
        ws["B8"] = f"คุณ : {cust_name}" if cust_name else "คุณ :"
        # B9: company
        ws["B9"] = cust_company
        # B13: email
        ws["B13"] = cust_email
        # B14: phone
        ws["B14"] = cust_phone

        # ===== ADD: keep only 1 sheet (print only this sheet) =====
        for name in list(wb.sheetnames):
            if name != TEMPLATE_SHEET_NAME:
                del wb[name]
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.print_area = "A1:I80"

        # Motor row (A20..G20)
        ws["A20"] = 1
        ws["B20"] = motor_code
        ws["C20"] = motor.get("desc", "")
        ws["F20"] = qty_motor
        ws["G20"] = motor.get("price", 0.0)

        # Gear row (A22..G22)
        ws["A22"] = 2
        ws["B22"] = gear_code
        ws["C22"] = gear.get("desc", "")
        ws["F22"] = qty_gear
        ws["G22"] = gear.get("price", 0.0)

        # Controller row (A24..G24) — optional
        if ctrl_model and qty_ctrl > 0:
            ws["A24"] = 3
            ws["B24"] = ctrl_model
            ws["C24"] = ctrl.get("desc", "")
            ws["F24"] = qty_ctrl
            ws["G24"] = ctrl.get("price", 0.0)

        # [BUG FIX #1, #3] เรียก next_qmo_no() และ cleanup_old_pdfs() เพียงครั้งเดียว
        # และ set เซลล์ H3, A17, D60, D61 เพียงครั้งเดียวที่นี่ (ไม่ซ้ำใน if block ข้างบน)
        cleanup_old_pdfs()

        run_no = next_qmo_no()
        run_no_str = f"{run_no:03d}"  # 001,002,003...

        # H3: QMO26-{abbr}-XXX — ใช้ชื่อย่อ Sale Person แทน SAS ถ้าเลือก Sale
        sp_abbr = sp.get("abbr", "").strip() if isinstance(sp, dict) else str(sp).strip()
        brand_tag = sp_abbr if sp_abbr else "SAS"
        ws["H3"] = f"QMO26-{brand_tag}-{run_no_str}"

        # A17: ชื่อย่อ Sale Person
        ws["A17"] = sp["abbr"]

        # D60: ชื่อเต็ม Sale Person
        ws["D60"] = sp["name"]

        # D61: ตำแหน่ง + เบอร์ Sale Person
        ws["D61"] = sp["position"]

        # Create filled xlsx and convert to pdf
        with tempfile.TemporaryDirectory() as td:
            filled_xlsx = os.path.join(td, "QMO26-SAS-FILLED.xlsx")
            wb.save(filled_xlsx)

            try:
                pdf_temp = xlsx_to_pdf(filled_xlsx, td)
            except Exception as e:
                return f"PDF convert failed (LibreOffice): {e}", 500

            # run_no_str ใช้ค่าที่ประกาศไว้แล้วข้างบน (ไม่ต้องเรียก next_qmo_no() อีก)

            company_for_file = cust_company or "NO-COMPANY"
            company_for_file = re.sub(r'[\\/:*?"<>|]+', "", company_for_file).strip()
            company_for_file = re.sub(r"\s+", "_", company_for_file)[:60] or "NO-COMPANY"

            date_str = datetime.now().strftime("%Y%m%d")  # วันที่
            saved_name = f"QMO26-{brand_tag}-{run_no_str}-{company_for_file}-{date_str}.pdf"
            saved_path = OUTPUT_DIR / saved_name
            shutil.copy2(pdf_temp, saved_path)

            # Optional: email to customer if SMTP configured
            customer = payload.get("customer", {}) or {}
            to_email = (customer.get("email") or "").strip()
            if smtp_is_configured():
                subject = f"SAS Quotation: {motor_code} + {gear_code}"

                # 1) ส่งให้ลูกค้า (ถ้ามีอีเมล)
                if to_email:
                    body_customer = (
                        f"เรียนคุณ {customer.get('name','')}\n\n"
                        f"ใบเสนอราคาของท่านถูกสร้างเรียบร้อยแล้ว\n"
                        f"Model: {payload.get('motorCode','')}\n"
                        f"Qty Motor: {qty_motor}\n"
                        f"Qty Gear: {qty_gear}\n\n"
                        f"แนบไฟล์ PDF มาพร้อมอีเมลนี้\n"
                    )
                    try:
                        send_email_with_pdf(to_email, subject, body_customer, str(saved_path))
                    except Exception as e:
                        # do not fail the download if email fails
                        print("[WARN] send_email_with_pdf (customer) failed:", e)

                # 2) ส่งสำเนาให้ทีม internal ทุกครั้งที่มีการดาวน์โหลดใบเสนอราคา
                body_internal = (
                    f"แจ้งเตือน: มีการดาวน์โหลดใบเสนอราคา AC Gear Motor\n\n"
                    f"ลูกค้า  : {customer.get('name', '-')}\n"
                    f"บริษัท  : {customer.get('company', '-')}\n"
                    f"เบอร์   : {customer.get('phone', '-')}\n"
                    f"อีเมล   : {customer.get('email', '-')}\n\n"
                    f"Model   : {payload.get('motorCode', '')}\n"
                    f"Motor   : {motor_code}  x{qty_motor}\n"
                    f"Gear    : {gear_code}   x{qty_gear}\n"
                    f"Sale    : {sp.get('name', sale_person_abbr)}\n"
                    f"เลขที่  : QMO26-{brand_tag}-{run_no_str}\n\n"
                    f"แนบไฟล์ PDF มาพร้อมอีเมลนี้\n"
                )
                for internal_email in INTERNAL_NOTIFY_EMAILS:
                    try:
                        send_email_with_pdf(internal_email, subject, body_internal, str(saved_path))
                    except Exception as e:
                        print(f"[WARN] send_email_with_pdf (internal {internal_email}) failed:", e)

            # [BUG FIX #4] แก้ indentation ของ return send_file ให้อยู่นอก for-loop
            # Keep compatibility with existing frontend (expects PDF as response)
            return send_file(
                str(saved_path),
                mimetype="application/pdf",
                as_attachment=True,
                download_name=saved_name,
            )

    except Exception as e:
        # ✅ จุดนี้ทำให้ "ไม่เห็น HTML 500 อีก" และจะรู้สาเหตุจริง 100%
        app.logger.exception("ac_quote crashed")
        return jsonify({"ok": False, "error": str(e)}), 500


# =========================
# BLDC Price Map
# =========================

def build_bldc_price_map() -> dict[str, dict[str, object]]:
    """อ่านราคาจาก High BLDC Gear Motor Price List xlsx — โครงสร้างเดียวกับ AC"""
    if not os.path.exists(BLDC_PRICE_FILE):
        raise FileNotFoundError(f"BLDC Price file not found: {BLDC_PRICE_FILE}")

    wb = load_workbook(BLDC_PRICE_FILE, data_only=True)
    ws = wb.active

    out: dict[str, dict[str, object]] = {}
    for r in range(1, (ws.max_row or 0) + 1):
        code = ws.cell(r, 1).value  # Col A = Code
        if not code:
            continue
        raw = str(code).strip()
        key = norm_code(raw)

        desc  = ws.cell(r, 2).value  # Col B = Detail
        price = ws.cell(r, 4).value  # Col D = Price

        try:
            price_val = float(price) if price not in (None, "") else 0.0
        except Exception:
            price_val = 0.0

        out[key] = {
            "raw":   raw,
            "desc":  str(desc).strip() if desc is not None else "",
            "price": price_val,
        }
    return out

try:
    BLDC_PRICE_MAP = build_bldc_price_map()
except Exception as _e:
    BLDC_PRICE_MAP = {}
    print(f"[WARN] Cannot load BLDC_PRICE_MAP: {_e}")

@app.get("/api/bldc-price-reload")
def bldc_price_reload():
    global BLDC_PRICE_MAP
    try:
        BLDC_PRICE_MAP = build_bldc_price_map()
        return jsonify({"ok": True, "count": len(BLDC_PRICE_MAP)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

# =========================
# BLDC Quote endpoint
# =========================

@app.route("/api/bldc-quote", methods=["OPTIONS"])
def bldc_quote_preflight():
    return ("", 204)

@app.post("/api/bldc-quote")
def bldc_quote():
    """
    สร้างใบเสนอราคา BLDC Gear Motor
    payload:
        modelCode   – รหัสรุ่นเต็ม  e.g. "Z6BLD400-220-GV-30S-6GU30V"
        qtyMotor    – จำนวน Motor (int)
        qtyDriver   – จำนวน Driver (int)
        customer    – { name, company, phone, email }
        category    – "BLDCGearmotor" | "HighefficiencyBLDCGearmotor"
        heType      – "S" | "SF" | "SL" | "" (สำหรับ HE series)
        salePerson  – "CA" (abbr)
    """
    try:
        payload = request.get_json(silent=True) or {}

        model_code = str(payload.get("modelCode", "")).strip()
        category   = str(payload.get("category",  "")).strip()
        he_type    = str(payload.get("heType",    "")).strip()

        try:
            qty_motor  = int(payload.get("qtyMotor",  1) or 1)
        except Exception:
            qty_motor = 1

        try:
            qty_driver = int(payload.get("qtyDriver", 0) or 0)
        except Exception:
            qty_driver = 0

        driver_model = str(payload.get("driverModel", "") or "").strip()

        if not model_code:
            return "modelCode is required", 400

        if not os.path.exists(TEMPLATE_FILE):
            return f"Template file not found: {TEMPLATE_FILE}", 500

        if not BLDC_PRICE_MAP:
            return "BLDC_PRICE_MAP is empty. Check BLDC_PRICE_FILE or reload /api/bldc-price-reload", 500

        # ── แยก motorCode และ gearCode จาก modelCode ────────────────────────────
        # HE series: ZxBLDppp-220-G??-SS-gearPart
        #   motorCode = ZxBLDppp-220G??-SS   (key ใน price list)
        #   gearCode  = gearPart (segment สุดท้าย)
        # Normal BLDC: ZxBLDppp-VV-GT-SS-gearPart
        #   ไม่มีราคา Motor แยกในไฟล์ BLDC — ราคาอยู่ใน ZxBLD...-220GV-SS เท่านั้น
        # ─────────────────────────────────────────────────────────────────────────

        parts = model_code.split("-")
        gear_code  = parts[-1] if parts else ""

        # สร้าง motor key สำหรับค้นหาใน BLDC_PRICE_MAP
        if "-220-" in model_code:
            # HE series: ZxBLDppp-220-G??-SS-... → ZxBLDppp-220G??-SS
            import re as _re
            m = _re.match(r"^(Z\dBLD\d+)-220-(G[A-Z]+)-(\d+S)", model_code, _re.IGNORECASE)
            motor_code = f"{m.group(1)}-220{m.group(2)}-{m.group(3)}" if m else ""
        else:
            # Normal BLDC (DC 24/36/48V): ราคา motor ค้นจาก ZxBLDppp-VV-GT-SS
            # ไฟล์ราคาใช้ format: Z5BLD90-24GU-30S (ไม่มี - ระหว่าง voltage กับ gear type)
            # model code จาก frontend: Z5BLD90-24-GU-30S-5GU3KB
            # parts[0]=Z5BLD90, [1]=24, [2]=GU, [3]=30S, [4]=gear
            motor_code = f"{parts[0]}-{parts[1]}{parts[2]}-{parts[3]}" if len(parts) >= 5 else ""

        mkey = norm_code(motor_code)
        gkey = norm_code(gear_code)

        motor = BLDC_PRICE_MAP.get(mkey)
        gear  = BLDC_PRICE_MAP.get(gkey)

        if not motor:
            return f"Motor code not found in BLDC price list: \"{motor_code}\" (key={mkey})", 404
        if not gear:
            return f"Gear code not found in BLDC price list: \"{gear_code}\" (key={gkey})", 404

        # Driver (optional) — ถ้าหาราคาไม่เจอในไฟล์ ใส่รหัสได้เลย ราคา = 0
        driver = None
        if driver_model and qty_driver > 0:
            dkey = norm_code(driver_model)
            driver = BLDC_PRICE_MAP.get(dkey) or {
                "raw":   driver_model,
                "desc":  f"BLDC Driver {driver_model}",
                "price": 0.0,
            }

        # ── Customer & Sale Person ────────────────────────────────────────────
        cust = payload.get("customer", {}) or {}
        cust_name    = str(cust.get("name",    "")).strip()
        cust_company = str(cust.get("company", "")).strip()
        cust_phone   = str(cust.get("phone",   "")).strip()
        cust_email   = str(cust.get("email",   "")).strip()

        sale_person_abbr = str(payload.get("salePerson", "CA")).strip() or "CA"
        sp = SALE_PERSONS.get(sale_person_abbr, SALE_PERSONS["CA"])

        # ── Fill template (QMO26-SAS.xlsx) เหมือน AC ────────────────────────
        wb_t = load_workbook(TEMPLATE_FILE)
        if TEMPLATE_SHEET_NAME not in wb_t.sheetnames:
            return f"Template sheet not found: {TEMPLATE_SHEET_NAME}", 500

        ws = wb_t[TEMPLATE_SHEET_NAME]

        # เก็บแค่ชีทที่ต้องการ
        for name in list(wb_t.sheetnames):
            if name != TEMPLATE_SHEET_NAME:
                del wb_t[name]
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.print_area = "A1:I80"

        # ข้อมูลลูกค้า
        ws["B8"]  = f"คุณ : {cust_name}" if cust_name else "คุณ :"
        ws["B9"]  = cust_company
        ws["B13"] = cust_email
        ws["B14"] = cust_phone

        # Motor row (A20..G20)
        ws["A20"] = 1
        ws["B20"] = motor_code
        ws["C20"] = motor.get("desc", "")
        ws["F20"] = qty_motor
        ws["G20"] = motor.get("price", 0.0)

        # Gear row (A22..G22)
        ws["A22"] = 2
        ws["B22"] = gear_code
        ws["C22"] = gear.get("desc", "")
        ws["F22"] = qty_motor   # gearhead qty ตาม motor
        ws["G22"] = gear.get("price", 0.0)

        # Driver row (A24..G24) — แสดงเฉพาะเมื่อเลือก driver
        if driver and driver_model and qty_driver > 0:
            ws["A24"] = 3
            ws["B24"] = driver_model
            ws["C24"] = driver.get("desc", "")
            ws["F24"] = qty_driver
            ws["G24"] = driver.get("price", 0.0)

        # Quotation number & Sale Person
        cleanup_old_pdfs()
        run_no = next_qmo_no()
        run_no_str = f"{run_no:03d}"

        sp_abbr   = sp.get("abbr", "").strip() if isinstance(sp, dict) else str(sp).strip()
        brand_tag = sp_abbr if sp_abbr else "SAS"

        ws["H3"]  = f"QMO26-{brand_tag}-{run_no_str}"
        ws["A17"] = sp["abbr"]
        ws["D60"] = sp["name"]
        ws["D61"] = sp["position"]
        ws["B17"] = "24 Months"
        ws["D17"] = "BLDC Gear Motor"
        ws["F17"] = "Within 7-14 days"

        # ── Convert to PDF ────────────────────────────────────────────────────
        with tempfile.TemporaryDirectory() as td:
            filled_xlsx = os.path.join(td, "QMO26-BLDC-FILLED.xlsx")
            wb_t.save(filled_xlsx)

            try:
                pdf_temp = xlsx_to_pdf(filled_xlsx, td)
            except Exception as e:
                return f"PDF convert failed (LibreOffice): {e}", 500

            company_for_file = cust_company or "NO-COMPANY"
            company_for_file = re.sub(r'[\\/:*?"<>|]+', "", company_for_file).strip()
            company_for_file = re.sub(r"\s+", "_", company_for_file)[:60] or "NO-COMPANY"

            date_str   = datetime.now().strftime("%Y%m%d")
            saved_name = f"QMO26-BLDC-{brand_tag}-{run_no_str}-{company_for_file}-{date_str}.pdf"
            saved_path = OUTPUT_DIR / saved_name
            shutil.copy2(pdf_temp, saved_path)

            # Optional SMTP email
            if smtp_is_configured():
                subject = f"SAS Quotation (BLDC): {model_code}"
                if cust_email:
                    body_customer = (
                        f"เรียนคุณ {cust_name}\n\n"
                        f"ใบเสนอราคา BLDC Gear Motor ของท่านถูกสร้างเรียบร้อยแล้ว\n"
                        f"Model: {model_code}\n"
                        f"จำนวน Motor: {qty_motor} ตัว | Driver: {qty_driver} ตัว\n\n"
                        f"แนบไฟล์ PDF มาพร้อมอีเมลนี้\n"
                    )
                    try:
                        send_email_with_pdf(cust_email, subject, body_customer, str(saved_path))
                    except Exception as e:
                        print("[WARN] bldc send_email_with_pdf (customer) failed:", e)

                body_internal = (
                    f"แจ้งเตือน: มีการดาวน์โหลดใบเสนอราคา BLDC Gear Motor\n\n"
                    f"ลูกค้า  : {cust_name}\n"
                    f"บริษัท  : {cust_company}\n"
                    f"เบอร์   : {cust_phone}\n"
                    f"อีเมล   : {cust_email}\n\n"
                    f"Model   : {model_code}\n"
                    f"Motor   : {motor_code}  x{qty_motor}\n"
                    f"Gear    : {gear_code}   x{qty_driver}\n"
                    f"Category: {category} {he_type}\n"
                    f"Sale    : {sp.get('name', sale_person_abbr)}\n"
                    f"เลขที่  : QMO26-BLDC-{brand_tag}-{run_no_str}\n\n"
                    f"แนบไฟล์ PDF มาพร้อมอีเมลนี้\n"
                )
                for internal_email in INTERNAL_NOTIFY_EMAILS:
                    try:
                        send_email_with_pdf(internal_email, subject, body_internal, str(saved_path))
                    except Exception as e:
                        print(f"[WARN] bldc send_email_with_pdf (internal {internal_email}) failed:", e)

            return send_file(
                str(saved_path),
                mimetype="application/pdf",
                as_attachment=True,
                download_name=saved_name,
            )

    except Exception as e:
        app.logger.exception("bldc_quote crashed")
        return jsonify({"ok": False, "error": str(e)}), 500

# =========================
# IEC Motor Quote endpoint
# =========================

@app.route("/api/iec-quote", methods=["OPTIONS"])
def iec_quote_preflight():
    return ("", 204)

@app.post("/api/iec-quote")
def iec_quote():
    """
    สร้างใบเสนอราคา IEC Standard Motor
    payload fields:
        modelCode       – รหัสรุ่นเต็ม  e.g. "YE3-112M-4-B5-0-X"
        iecMotorType    – "YE3" / "YE4" / "YVP" / "YEJ" / "YPEJ" / "YB"
        iecPower        – "4" (kW string)
        iecPole         – "4P"
        iecMount        – "B5"
        iecTerminal     – "0"
        iecCable        – "X"
        qtyMotor        – จำนวน (int)
        unitPrice       – ราคาต่อหน่วย (float, คำนวณจาก frontend)
        customer        – { name, company, phone, email }
        salePerson      – "CA" (abbr)
    """
    try:
        APP_VERSION = "2026-IEC-v1"
        app.logger.info("iec_quote called, APP_VERSION=%s", APP_VERSION)

        payload = request.get_json(silent=True) or {}

        model_code     = str(payload.get("modelCode",    "")).strip()
        iec_motor_type = str(payload.get("iecMotorType", "")).strip()
        iec_power      = str(payload.get("iecPower",     "")).strip()
        iec_pole       = str(payload.get("iecPole",      "")).strip()
        iec_mount      = str(payload.get("iecMount",     "")).strip()
        iec_terminal   = str(payload.get("iecTerminal",  "")).strip()
        iec_cable      = str(payload.get("iecCable",     "")).strip()

        try:
            qty_motor = int(payload.get("qtyMotor", 1) or 1)
        except Exception:
            qty_motor = 1

        try:
            unit_price = float(payload.get("unitPrice", 0) or 0)
        except Exception:
            unit_price = 0.0

        total_price = unit_price * qty_motor

        if not model_code:
            return "Invalid modelCode", 400

        if not os.path.exists(TEMPLATE_FILE):
            return f"Template file not found: {TEMPLATE_FILE}", 500

        # ── Customer ──────────────────────────────────────────────────────────
        cust         = payload.get("customer", {}) or {}
        cust_name    = str(cust.get("name",    "")).strip()
        cust_company = str(cust.get("company", "")).strip()
        cust_phone   = str(cust.get("phone",   "")).strip()
        cust_email   = str(cust.get("email",   "")).strip()

        # ── Sale Person ───────────────────────────────────────────────────────
        sale_person_abbr = str(payload.get("salePerson", "CA")).strip() or "CA"
        sp = SALE_PERSONS.get(sale_person_abbr, SALE_PERSONS["CA"])

        # ── Load template ─────────────────────────────────────────────────────
        wb = load_workbook(TEMPLATE_FILE)
        if TEMPLATE_SHEET_NAME not in wb.sheetnames:
            return f"Template sheet not found: {TEMPLATE_SHEET_NAME}", 500
        ws = wb[TEMPLATE_SHEET_NAME]

        # ── Keep only this sheet ──────────────────────────────────────────────
        for name in list(wb.sheetnames):
            if name != TEMPLATE_SHEET_NAME:
                del wb[name]
        ws.page_setup.paperSize  = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.print_area = "A1:I80"

        # ── Customer cells ────────────────────────────────────────────────────
        ws["B8"]  = f"คุณ : {cust_name}" if cust_name else "คุณ :"
        ws["B9"]  = cust_company
        ws["B13"] = cust_email
        ws["B14"] = cust_phone

        # ── Quotation number ──────────────────────────────────────────────────
        cleanup_old_pdfs()
        run_no     = next_qmo_no()
        run_no_str = f"{run_no:03d}"
        sp_abbr    = sp.get("abbr", "").strip() if isinstance(sp, dict) else str(sp).strip()
        brand_tag  = sp_abbr if sp_abbr else "SAS"
        ws["H3"]   = f"QMO26-{brand_tag}-{run_no_str}"

        # ── Sale Person cells ─────────────────────────────────────────────────
        ws["A17"] = sp["abbr"]
        ws["D60"] = sp["name"]
        ws["D61"] = sp["position"]

        # ── IEC Motor line item (row 20) ──────────────────────────────────────
        # ── Shaft Diameter lookup (ใช้ frame key เดียวกับ JS) ────────────────
        def get_shaft_diameter(model_code: str, iec_pole: str) -> str:
            """คืนค่า Ø mm จาก model_code + pole โดย normalize เป็น YE3 frame key"""
            # แยก frame จาก model เช่น YE3-100L1-4-B5-0-X → 100L1, pole → 4
            parts = model_code.split("-")
            # parts[0]=YE3/YE4/..., parts[1]=frame, parts[2]=pole
            if len(parts) < 3:
                return ""
            frame = parts[1]
            pole_num = iec_pole.replace("P", "").replace("p", "")  # '4P' → '4'
            ye3_key = f"YE3-{frame}-{pole_num}"

            # SHAFT_DIAMETER_MAP — 2P
            shaft_2p = {
                "YE3-80M1-2": 19, "YE3-80M2-2": 19,
                "YE3-90S-2":  24, "YE3-90L-2":  24,
                "YE3-100L-2": 28,
                "YE3-112M-2": 28,
                "YE3-132S1-2": 38, "YE3-132S2-2": 38,
                "YE3-160M1-2": 42, "YE3-160M2-2": 42, "YE3-160L-2": 42,
                "YE3-180M-2": 48,
                "YE3-200L1-2": 55, "YE3-200L2-2": 55, "YE3-225M-2": 55,
                "YE3-250M-2": 60,
                "YE3-280S-2": 65, "YE3-280M-2": 65,
                "YE3-315S-2": 65, "YE3-315M-2": 65, "YE3-315L1-2": 65, "YE3-315L2-2": 65,
            }
            # SHAFT_DIAMETER_MAP — 4P (6P/8P ใช้ค่าเดียวกัน)
            shaft_4p = {
                "YE3-71M10-4": 14, "YE3-71M11-4": 14, "YE3-71M22-4": 14,
                "YE3-80M1-4":  19, "YE3-80M2-4":  19,
                "YE3-90S-4":   24, "YE3-90L-4":   24,
                "YE3-100L1-4": 28, "YE3-100L2-4": 28,
                "YE3-112M-4":  28,
                "YE3-132S-4":  38, "YE3-132M-4":  38,
                "YE3-160M-4":  42, "YE3-160L-4":  42,
                "YE3-180M-4":  48, "YE3-180L-4":  48,
                "YE3-200L-4":  55,
                "YE3-225S-4":  60, "YE3-225M-4":  60,
                "YE3-250M-4":  65,
                "YE3-280S-4":  75, "YE3-280M-4":  75,
                "YE3-315S-4":  80, "YE3-315M-4":  80, "YE3-315L1-4": 80, "YE3-315L2-4": 80,
                "YE3-355M-4":  95, "YE3-355L-4":  95,
            }
            if pole_num == "2":
                val = shaft_2p.get(ye3_key)
            else:
                # 4P, 6P, 8P → normalize เป็น 4P
                key4 = f"YE3-{frame}-4"
                val = shaft_4p.get(key4)
            return f"\u00d8{val} mm" if val else ""

        shaft_str = get_shaft_diameter(model_code, iec_pole)

        pole_label = {
            "2P": "2 Pole (~3000 rpm)", "4P": "4 Pole (~1500 rpm)",
            "6P": "6 Pole (~1000 rpm)", "8P": "8 Pole (~750 rpm)"
        }.get(iec_pole, iec_pole)

        # ── ดึง spec จริงจาก YE3_DB (Python version) ─────────────────────────
        # key format เดียวกับ JS: YE3-{frame}-{pole_num}
        _pole_num = iec_pole.replace("P", "").replace("p", "")
        _parts    = model_code.split("-")
        _frame    = _parts[1] if len(_parts) > 1 else ""
        _ye3_key  = f"YE3-{_frame}-{_pole_num}"

        YE3_SPEC = {
            "YE3-80M1-2":  dict(speed=2880, eff=80.7, i380=1.7,  i400=1.6,  i415=1.6,  torque=2.49,  weight=18.1),
            "YE3-80M2-2":  dict(speed=2880, eff=82.7, i380=2.4,  i400=2.3,  i415=2.2,  torque=3.65,  weight=19.5),
            "YE3-90S-2":   dict(speed=2895, eff=84.2, i380=3.2,  i400=3.1,  i415=3.0,  torque=4.95,  weight=23.3),
            "YE3-90L-2":   dict(speed=2895, eff=85.9, i380=4.6,  i400=4.3,  i415=4.2,  torque=7.26,  weight=27.1),
            "YE3-100L-2":  dict(speed=2895, eff=87.1, i380=6.0,  i400=5.7,  i415=5.5,  torque=9.9,   weight=38.8),
            "YE3-112M-2":  dict(speed=2905, eff=88.1, i380=7.8,  i400=7.4,  i415=7.2,  torque=13.1,  weight=48.3),
            "YE3-132S1-2": dict(speed=2930, eff=89.2, i380=10.6, i400=10.1, i415=9.7,  torque=17.9,  weight=55.1),
            "YE3-132S2-2": dict(speed=2930, eff=90.1, i380=14.4, i400=13.7, i415=13.2, torque=24.4,  weight=69.2),
            "YE3-160M1-2": dict(speed=2945, eff=91.2, i380=20.6, i400=19.6, i415=18.9, torque=35.7,  weight=113),
            "YE3-160M2-2": dict(speed=2945, eff=91.9, i380=27.9, i400=26.5, i415=25.5, torque=48.6,  weight=123),
            "YE3-160L-2":  dict(speed=2940, eff=92.4, i380=34.2, i400=32.5, i415=31.3, torque=60.1,  weight=142),
            "YE3-180M-2":  dict(speed=2955, eff=92.7, i380=40.5, i400=38.5, i415=37.1, torque=71.1,  weight=182),
            "YE3-200L1-2": dict(speed=2960, eff=93.3, i380=54.9, i400=52.1, i415=50.3, torque=96.8,  weight=246),
            "YE3-200L2-2": dict(speed=2960, eff=93.7, i380=67.4, i400=64.0, i415=61.7, torque=119.4, weight=265),
            "YE3-225M-2":  dict(speed=2965, eff=94.0, i380=80.8, i400=76.8, i415=74.0, torque=144.9, weight=323),
            "YE3-250M-2":  dict(speed=2970, eff=94.3, i380=98.5, i400=93.5, i415=90.2, torque=176.9, weight=413),
            "YE3-280S-2":  dict(speed=2975, eff=94.7, i380=133.7,i400=127.0,i415=122.4,torque=240.8, weight=546),
            "YE3-280M-2":  dict(speed=2975, eff=95.0, i380=159.9,i400=151.9,i415=146.4,torque=288.9, weight=569),
            "YE3-315S-2":  dict(speed=2975, eff=95.2, i380=195.1,i400=185.3,i415=178.6,torque=352.8, weight=897),
            "YE3-315M-2":  dict(speed=2978, eff=95.4, i380=233.6,i400=221.9,i415=213.9,torque=423.3, weight=1029),
            "YE3-315L1-2": dict(speed=2978, eff=95.6, i380=279.4,i400=265.5,i415=255.9,torque=512.8, weight=1067),
            "YE3-315L2-2": dict(speed=2980, eff=95.8, i380=348.6,i400=331.1,i415=319.2,torque=640.9, weight=1194),
            "YE3-71M10-4": dict(speed=1330, eff=69.9, i380=0.5,  i400=0.51, i415=0.49, torque=1.29,  weight=4.5),
            "YE3-71M11-4": dict(speed=1350, eff=73.5, i380=0.7,  i400=0.66, i415=0.64, torque=1.77,  weight=6.5),
            "YE3-71M22-4": dict(speed=1350, eff=77.3, i380=1.0,  i400=0.92, i415=0.89, torque=2.62,  weight=7.5),
            "YE3-80M1-4":  dict(speed=1400, eff=80.8, i380=1.4,  i400=1.3,  i415=1.3,  torque=3.75,  weight=17.6),
            "YE3-80M2-4":  dict(speed=1420, eff=82.5, i380=1.8,  i400=1.7,  i415=1.7,  torque=5.04,  weight=18.4),
            "YE3-90S-4":   dict(speed=1445, eff=84.1, i380=2.6,  i400=2.5,  i415=2.4,  torque=7.27,  weight=24.2),
            "YE3-90L-4":   dict(speed=1445, eff=85.3, i380=3.5,  i400=3.3,  i415=3.2,  torque=9.91,  weight=29.7),
            "YE3-100L1-4": dict(speed=1435, eff=86.7, i380=4.8,  i400=4.5,  i415=4.4,  torque=14.6,  weight=41.5),
            "YE3-100L2-4": dict(speed=1435, eff=87.7, i380=6.3,  i400=6.0,  i415=5.8,  torque=20.0,  weight=46),
            "YE3-112M-4":  dict(speed=1440, eff=88.6, i380=8.4,  i400=7.9,  i415=7.7,  torque=26.5,  weight=63.2),
            "YE3-132S-4":  dict(speed=1460, eff=89.6, i380=11.2, i400=10.7, i415=10.3, torque=36.0,  weight=71.2),
            "YE3-132M-4":  dict(speed=1460, eff=90.4, i380=15.0, i400=14.3, i415=13.7, torque=49.1,  weight=85.1),
            "YE3-160M-4":  dict(speed=1465, eff=91.4, i380=21.5, i400=20.4, i415=19.7, torque=71.7,  weight=121),
            "YE3-160L-4":  dict(speed=1465, eff=92.1, i380=28.8, i400=27.3, i415=26.3, torque=97.8,  weight=142),
            "YE3-180M-4":  dict(speed=1470, eff=92.6, i380=35.3, i400=33.5, i415=32.3, torque=120.2, weight=181),
            "YE3-180L-4":  dict(speed=1470, eff=93.0, i380=41.8, i400=39.7, i415=38.3, torque=142.9, weight=209),
            "YE3-200L-4":  dict(speed=1475, eff=93.6, i380=56.6, i400=53.8, i415=51.9, torque=194.2, weight=284),
            "YE3-225S-4":  dict(speed=1485, eff=93.9, i380=69.6, i400=66.1, i415=63.7, torque=237.9, weight=328),
            "YE3-225M-4":  dict(speed=1485, eff=94.2, i380=84.4, i400=80.2, i415=77.3, torque=289.4, weight=363),
            "YE3-250M-4":  dict(speed=1485, eff=94.6, i380=102.7,i400=97.6, i415=94.1, torque=353.7, weight=442),
            "YE3-280S-4":  dict(speed=1486, eff=95.0, i380=136.3,i400=129.5,i415=124.8,torque=482.0, weight=569),
            "YE3-280M-4":  dict(speed=1486, eff=95.2, i380=163.2,i400=155.1,i415=149.5,torque=578.4, weight=639),
            "YE3-315S-4":  dict(speed=1488, eff=95.4, i380=196.8,i400=187.0,i415=180.2,torque=706.0, weight=939),
            "YE3-315M-4":  dict(speed=1488, eff=95.6, i380=235.7,i400=223.9,i415=215.8,torque=847.2, weight=1033),
            "YE3-315L1-4": dict(speed=1488, eff=95.8, i380=285.1,i400=270.9,i415=261.1,torque=1027,  weight=1126),
            "YE3-315L2-4": dict(speed=1490, eff=96.0, i380=351.7,i400=334.1,i415=322.0,torque=1282,  weight=1238),
            "YE3-355M-4":  dict(speed=1490, eff=96.0, i380=439.6,i400=417.6,i415=402.6,torque=1602,  weight=1830),
            "YE3-355L-4":  dict(speed=1490, eff=96.0, i380=553.9,i400=526.2,i415=507.2,torque=2019,  weight=1950),
            "YE3-80M2-6":  dict(speed=890,  eff=73.6, i380=1.6,  i400=1.5,  i415=1.4,  torque=5.9,   weight=16.6),
            "YE3-90S-6":   dict(speed=935,  eff=78.9, i380=2.0,  i400=1.9,  i415=1.8,  torque=7.66,  weight=24.1),
            "YE3-90L-6":   dict(speed=945,  eff=81.0, i380=2.8,  i400=2.7,  i415=2.6,  torque=11.1,  weight=25.7),
            "YE3-100L-6":  dict(speed=949,  eff=82.5, i380=3.7,  i400=3.5,  i415=3.4,  torque=15.1,  weight=34.9),
            "YE3-112M-6":  dict(speed=955,  eff=84.3, i380=5.4,  i400=5.1,  i415=4.9,  torque=22.0,  weight=54.2),
            "YE3-132S-6":  dict(speed=968,  eff=85.6, i380=7.2,  i400=6.8,  i415=6.6,  torque=29.6,  weight=62.3),
            "YE3-132M1-6": dict(speed=968,  eff=86.8, i380=9.5,  i400=9.0,  i415=8.7,  torque=39.5,  weight=75.2),
            "YE3-132M2-6": dict(speed=968,  eff=88.0, i380=12.7, i400=12.0, i415=11.6, torque=54.3,  weight=82.3),
            "YE3-160M-6":  dict(speed=970,  eff=89.1, i380=16.2, i400=15.4, i415=14.8, torque=73.8,  weight=112),
            "YE3-160L-6":  dict(speed=970,  eff=90.3, i380=23.1, i400=22.0, i415=21.2, torque=108.3, weight=134),
            "YE3-180L-6":  dict(speed=978,  eff=91.2, i380=30.9, i400=29.3, i415=28.2, torque=146.5, weight=197),
            "YE3-200L1-6": dict(speed=980,  eff=91.7, i380=37.8, i400=36.0, i415=34.7, torque=180.3, weight=234),
            "YE3-200L2-6": dict(speed=980,  eff=92.2, i380=44.8, i400=42.5, i415=41.0, torque=214.4, weight=251),
            "YE3-225M-6":  dict(speed=980,  eff=92.9, i380=59.1, i400=56.2, i415=54.1, torque=292.3, weight=308),
            "YE3-250M-6":  dict(speed=985,  eff=93.3, i380=71.7, i400=68.1, i415=65.7, torque=358.7, weight=383),
            "YE3-280S-6":  dict(speed=985,  eff=93.7, i380=85.8, i400=81.6, i415=78.6, torque=436.3, weight=501),
            "YE3-280M-6":  dict(speed=985,  eff=94.1, i380=103.3,i400=98.1, i415=94.6, torque=533.2, weight=573),
            "YE3-315S-6":  dict(speed=985,  eff=94.6, i380=143.4,i400=136.2,i415=131.3,torque=727.2, weight=843),
            "YE3-315M-6":  dict(speed=988,  eff=94.9, i380=169.5,i400=161.0,i415=155.2,torque=869.9, weight=941),
            "YE3-315L1-6": dict(speed=988,  eff=95.1, i380=206.8,i400=196.4,i415=189.3,torque=1063,  weight=1017),
            "YE3-315L2-6": dict(speed=988,  eff=95.4, i380=244.5,i400=232.2,i415=223.8,torque=1276,  weight=1121),
            "YE3-80M1-8":  dict(speed=650,  eff=51.0, i380=0.9,  i400=0.8,  i415=0.8,  torque=2.64,  weight=14.5),
            "YE3-90S-8":   dict(speed=675,  eff=62.0, i380=1.5,  i400=1.4,  i415=1.4,  torque=5.23,  weight=28.2),
            "YE3-90L-8":   dict(speed=675,  eff=63.0, i380=2.2,  i400=2.1,  i415=2.0,  torque=7.78,  weight=29.7),
            "YE3-100L1-8": dict(speed=685,  eff=68.7, i380=2.5,  i400=2.4,  i415=2.3,  torque=10.5,  weight=40),
            "YE3-100L2-8": dict(speed=685,  eff=70.7, i380=3.5,  i400=3.4,  i415=3.2,  torque=15.3,  weight=41.4),
            "YE3-112M-8":  dict(speed=695,  eff=72.8, i380=4.4,  i400=4.2,  i415=4.0,  torque=20.6,  weight=57.5),
            "YE3-132S-8":  dict(speed=710,  eff=77.9, i380=6.0,  i400=5.7,  i415=5.5,  torque=29.6,  weight=74.8),
            "YE3-132M-8":  dict(speed=710,  eff=78.9, i380=7.9,  i400=7.5,  i415=7.2,  torque=40.4,  weight=89.1),
            "YE3-160M1-8": dict(speed=725,  eff=79.9, i380=10.4, i400=9.9,  i415=9.5,  torque=52.7,  weight=101),
            "YE3-160M2-8": dict(speed=725,  eff=82.0, i380=13.8, i400=13.1, i415=12.6, torque=72.4,  weight=126.5),
            "YE3-160L-8":  dict(speed=725,  eff=84.0, i380=18.1, i400=17.2, i415=16.6, torque=98.8,  weight=136),
            "YE3-180L-8":  dict(speed=735,  eff=86.4, i380=25.8, i400=24.5, i415=23.6, torque=142.9, weight=198),
            "YE3-200L-8":  dict(speed=730,  eff=86.9, i380=34.5, i400=32.8, i415=31.6, torque=196.2, weight=234),
            "YE3-225S-8":  dict(speed=730,  eff=89.1, i380=41.5, i400=39.4, i415=38.0, torque=242.0, weight=284),
            "YE3-225M-8":  dict(speed=730,  eff=89.6, i380=47.8, i400=45.4, i415=43.8, torque=287.8, weight=325),
            "YE3-250M-8":  dict(speed=735,  eff=90.4, i380=63.8, i400=60.6, i415=58.4, torque=389.8, weight=425),
            "YE3-280S-8":  dict(speed=735,  eff=90.9, i380=78.3, i400=74.4, i415=71.7, torque=480.7, weight=518),
            "YE3-280M-8":  dict(speed=735,  eff=91.4, i380=94.7, i400=90.0, i415=86.7, torque=584.7, weight=582),
            "YE3-315S-8":  dict(speed=735,  eff=92.3, i380=113.2,i400=107.5,i415=103.6,torque=714.6, weight=852),
            "YE3-315M-8":  dict(speed=735,  eff=93.2, i380=152.8,i400=145.2,i415=139.9,torque=974.5, weight=952),
            "YE3-315L1-8": dict(speed=735,  eff=93.5, i380=182.8,i400=173.7,i415=167.4,torque=1169,  weight=1040),
            "YE3-315L2-8": dict(speed=735,  eff=93.5, i380=218.0,i400=207.1,i415=199.6,torque=1429,  weight=1056),
        }
        spec = YE3_SPEC.get(_ye3_key, {})

        ws["A20"] = 1
        ws["B20"] = model_code
        ws["C20"] = f"Power: {iec_power} kW | {iec_pole.replace('P',' Pole')} | Mounting: {iec_mount} | Terminal Box: {iec_terminal}\u00b0 | Cable: {iec_cable}"
        ws["C21"] = f"Standard: IEC 60034 / GB18613, 3Ph 380V 50Hz, IP55 | Motor Type: {iec_motor_type}"
        # ── Row 22 = แถว Gear ใน template (มี border) → ใส่ C22 ต่อ ─────────
        ws["C22"] = f"Insulation: Class F (155\u00b0C) | Cooling: IC411 | Duty: S1"
        ws["C23"] = f"Voltage: 200\u2013660V, 50/60Hz (\u00b15%) | Site: -15\u00b0C to +40\u00b0C, Alt \u2264 1000 m | Vibration: Class A (Class B on request)"
        # ── Row 24 = แถว Controller ใน template (มี border) → ใส่ C24 ต่อ ──
        ws["C24"] = f"Rated Speed: {spec['speed']} rpm" if spec.get('speed') else "Rated Speed: (see datasheet)"
        ws["C25"] = (f"Efficiency: {spec['eff']} %   |   Current: {spec['i380']} A @ 380V  /  {spec['i400']} A @ 400V  /  {spec['i415']} A @ 415V"
                     if spec.get('eff') and spec.get('i380') else "Efficiency / Current: (see datasheet)")
        ws["C26"] = f"Output Shaft: {shaft_str}" if shaft_str else "Output Shaft: (see datasheet)"
        ws["C27"] = f"Weight: {spec['weight']} kg" if spec.get('weight') else "Weight: (see datasheet)"
        ws["F20"] = qty_motor
        ws["G20"] = unit_price if unit_price > 0 else ""

        # ── Delivery date (F17) ───────────────────────────────────────────────
        ws["F17"] = "Within 30-45 days"

        # ── Warranty (B17) และ JOB (D17) ─────────────────────────────────────
        ws["B17"] = "24 Months"
        ws["D17"] = "IEC Standard Motor"

        # ── Convert to PDF ────────────────────────────────────────────────────
        with tempfile.TemporaryDirectory() as td:
            filled_xlsx = os.path.join(td, "QMO26-IEC-FILLED.xlsx")
            wb.save(filled_xlsx)

            try:
                pdf_temp = xlsx_to_pdf(filled_xlsx, td)
            except Exception as e:
                return f"PDF convert failed (LibreOffice): {e}", 500

            company_for_file = cust_company or "NO-COMPANY"
            company_for_file = re.sub(r'[\\/:*?"<>|]+', "", company_for_file).strip()
            company_for_file = re.sub(r"\s+", "_", company_for_file)[:60] or "NO-COMPANY"

            date_str   = datetime.now().strftime("%Y%m%d")
            saved_name = f"QMO26-{brand_tag}-{run_no_str}-{company_for_file}-{date_str}.pdf"
            saved_path = OUTPUT_DIR / saved_name
            shutil.copy2(pdf_temp, saved_path)

            # ── Optional SMTP email ───────────────────────────────────────────
            to_email = cust_email
            if smtp_is_configured():
                subject = f"SAS Quotation (IEC Motor): {model_code}"
                body_customer = (
                    f"เรียนคุณ {cust_name}\n\n"
                    f"ใบเสนอราคามอเตอร์ IEC ของท่านถูกสร้างเรียบร้อยแล้ว\n"
                    f"Model: {model_code}\n"
                    f"กำลัง: {iec_power} kW | Pole: {iec_pole} | Mounting: {iec_mount}\n"
                    f"จำนวน: {qty_motor} ตัว"
                    + (f" | ราคา/ตัว: {unit_price:,.0f} ฿ | รวม: {total_price:,.0f} ฿" if unit_price > 0 else "")
                    + "\n\nแนบไฟล์ PDF มาพร้อมอีเมลนี้\n"
                )
                if to_email:
                    try:
                        send_email_with_pdf(to_email, subject, body_customer, str(saved_path))
                    except Exception as e:
                        print("[WARN] send_email_with_pdf (customer) failed:", e)

                body_internal = (
                    f"แจ้งเตือน: มีการดาวน์โหลดใบเสนอราคา IEC Motor\n\n"
                    f"ลูกค้า  : {cust_name}\n"
                    f"บริษัท  : {cust_company}\n"
                    f"เบอร์   : {cust_phone}\n"
                    f"อีเมล   : {cust_email}\n\n"
                    f"Model   : {model_code}\n"
                    f"ประเภท  : {iec_motor_type} | {iec_power} kW | {iec_pole} | {iec_mount}\n"
                    f"จำนวน   : {qty_motor} ตัว"
                    + (f" | ราคา/ตัว: {unit_price:,.0f} ฿ | รวม: {total_price:,.0f} ฿" if unit_price > 0 else "")
                    + f"\nSale    : {sp.get('name', sale_person_abbr)}\n"
                    f"เลขที่  : QMO26-{brand_tag}-{run_no_str}\n\n"
                    f"แนบไฟล์ PDF มาพร้อมอีเมลนี้\n"
                )
                for internal_email in INTERNAL_NOTIFY_EMAILS:
                    try:
                        send_email_with_pdf(internal_email, subject, body_internal, str(saved_path))
                    except Exception as e:
                        print(f"[WARN] send_email_with_pdf (internal {internal_email}) failed:", e)

            return send_file(
                str(saved_path),
                mimetype="application/pdf",
                as_attachment=True,
                download_name=saved_name,
            )

    except Exception as e:
        app.logger.exception("iec_quote crashed")
        return jsonify({"ok": False, "error": str(e)}), 500

# =========================
# Mr.Motor Bot — LINE Webhook
# =========================
#
# ENV vars ที่ต้องตั้งใน Render:
#   LINE_CHANNEL_SECRET        — Channel Secret จาก LINE Developers Console
#   LINE_CHANNEL_ACCESS_TOKEN  — Long-lived Access Token
#   R2_PUBLIC_BASE             — https://pub-8cdc08b3fc55463c8c8f399a10351d7e.r2.dev
#
# คำสั่งที่รองรับ (ในกลุ่ม LINE):
#   ขอสเปค Model : 5IK90RGU-CF-5GU15KB   → ส่ง link PDF กลับ
#   ขอ3D Model : 5IK90RGU-CF-5GU15KB     → ส่ง link .STEP กลับ (normalize ratio→3)
# =========================

_LINE_REPLY_API = "https://api.line.me/v2/bot/message/reply"
_R2_BASE        = os.environ.get("R2_PUBLIC_BASE",
                    "https://pub-8cdc08b3fc55463c8c8f399a10351d7e.r2.dev")
_LINE_SECRET    = os.environ.get("LINE_CHANNEL_SECRET", "")
_LINE_TOKEN     = os.environ.get("LINE_CHANNEL_ACCESS_TOKEN", "")

_CMD_SPEC = re.compile(r"ขอสเปค\s*[Mm]odel\s*:\s*(.+)", re.IGNORECASE)
_CMD_3D   = re.compile(r"ขอ3[Dd]\s*[Mm]odel\s*:\s*(.+)", re.IGNORECASE)


def _normalize_ac_ratio(model_code: str) -> str:
    """เปลี่ยน ratio ทุกค่า → 3 เพื่อชี้ไฟล์จริงใน R2
    เช่น 5IK90RGU-CF-5GU15KB → 5IK90RGU-CF-5GU3KB"""
    return re.sub(
        r"(\d+)(GN|GU)(\d+(?:\.\d+)?)(KB|RC|RT|K)(\b|$)",
        lambda m: f"{m.group(1)}{m.group(2).upper()}3{m.group(4).upper()}",
        model_code,
        flags=re.IGNORECASE,
    )


def _verify_line_sig(body: bytes, signature: str) -> bool:
    """ตรวจ X-Line-Signature ป้องกัน request ปลอม"""
    if not _LINE_SECRET:
        return True  # dev mode — ข้ามการตรวจ
    mac = hmac.new(_LINE_SECRET.encode(), body, hashlib.sha256).digest()
    return hmac.compare_digest(base64.b64encode(mac).decode(), signature)


def _r2_url(filename: str) -> str:
    from urllib.parse import quote
    return f"{_R2_BASE}/{quote(filename, safe='')}"


def _reply_flex(reply_token: str, title: str, model: str,
                dl_url: str, btn_label: str, color: str):
    """ส่ง Flex Message พร้อมปุ่มดาวน์โหลด"""
    flex_body = {
        "type": "bubble",
        "size": "mega",
        "header": {
            "type": "box",
            "layout": "vertical",
            "backgroundColor": color,
            "paddingAll": "16px",
            "contents": [{
                "type": "text",
                "text": title,
                "color": "#FFFFFF",
                "size": "lg",
                "weight": "bold"
            }]
        },
        "body": {
            "type": "box",
            "layout": "vertical",
            "spacing": "sm",
            "paddingAll": "16px",
            "contents": [
                {
                    "type": "text",
                    "text": "Model",
                    "color": "#888888",
                    "size": "sm"
                },
                {
                    "type": "text",
                    "text": model,
                    "color": "#111111",
                    "size": "md",
                    "weight": "bold",
                    "wrap": True
                }
            ]
        },
        "footer": {
            "type": "box",
            "layout": "vertical",
            "paddingAll": "12px",
            "contents": [{
                "type": "button",
                "style": "primary",
                "color": color,
                "height": "sm",
                "action": {
                    "type": "uri",
                    "label": btn_label,
                    "uri": dl_url
                }
            }]
        }
    }

    http_requests.post(
        _LINE_REPLY_API,
        headers={
            "Authorization": f"Bearer {_LINE_TOKEN}",
            "Content-Type": "application/json",
        },
        json={
            "replyToken": reply_token,
            "messages": [{
                "type": "flex",
                "altText": f"{title} — {model}",
                "contents": flex_body
            }]
        },
        timeout=10,
    )



def _fetch_r2(filename: str) -> bytes | None:
    """ดึงไฟล์จาก R2 คืน bytes หรือ None ถ้าไม่เจอ/ไม่ใช่ไฟล์จริง"""
    try:
        r = http_requests.get(_r2_url(filename), timeout=30)
        if r.status_code != 200:
            return None
        if "text/html" in r.headers.get("content-type", ""):
            return None
        if len(r.content) < 256:
            return None
        return r.content
    except Exception as e:
        print(f"[R2 fetch error] {e}")
        return None


def _reply_text(reply_token: str, text: str):
    """ส่ง text message กลับ LINE group"""
    http_requests.post(
        _LINE_REPLY_API,
        headers={
            "Authorization": f"Bearer {_LINE_TOKEN}",
            "Content-Type": "application/json",
        },
        json={
            "replyToken": reply_token,
            "messages": [{"type": "text", "text": text}],
        },
        timeout=10,
    )


@app.post("/line/webhook")
def line_webhook():
    body_bytes = request.get_data()
    payload = request.get_json(silent=True) or {}

    for event in payload.get("events", []):
        if event.get("type") != "message":
            continue
        msg = event.get("message", {})
        if msg.get("type") != "text":
            continue

        text        = msg.get("text", "").strip()
        reply_token  = event.get("replyToken", "")
        _SERVER_BASE = "https://gear-motor-app.onrender.com"

        # ── คำสั่ง: ขอสเปค ───────────────────────────────────
        m = _CMD_SPEC.match(text)
        if m:
            raw_model   = m.group(1).strip()
            pdf_r2      = f"{raw_model}.pdf"
            pdf_display = f"{raw_model}.pdf"
            from urllib.parse import quote
            dl_url = (f"{_SERVER_BASE}/line/download"
                      f"?file={quote(pdf_r2, safe='')}"
                      f"&name={quote(pdf_display, safe='')}")
            _reply_flex(reply_token,
                title="📄 Spec Sheet",
                model=raw_model,
                dl_url=dl_url,
                btn_label="📥 ดาวน์โหลด PDF",
                color="#1DB954",
            )
            continue

        # ── คำสั่ง: ขอ3D ─────────────────────────────────────
        m = _CMD_3D.match(text)
        if m:
            raw_model    = m.group(1).strip()
            file_model   = _normalize_ac_ratio(raw_model)
            step_r2      = f"{file_model}.STEP"
            step_display = f"{raw_model}.STEP"
            from urllib.parse import quote
            dl_url = (f"{_SERVER_BASE}/line/download"
                      f"?file={quote(step_r2, safe='')}"
                      f"&name={quote(step_display, safe='')}")
            _reply_flex(reply_token,
                title="📦 3D Model",
                model=raw_model,
                dl_url=dl_url,
                btn_label="📥 ดาวน์โหลด .STEP",
                color="#0066CC",
            )
            continue

        # ── ข้อความอื่น ๆ ────────────────────────────────────
        _reply_text(reply_token,
            "สวัสดีครับ 😊 ผม Mr.SAS MotorBot\n\n"
            "ตอนนี้ผมเก่งแค่เรื่อง Spec กับ 3D ครับ\n"
            "เรื่องอื่นๆ อาจจะยังน้าาาา 🙏\n\n"
            "📌 คำสั่งที่ใช้ได้:\n"
            "• ขอสเปค Model : [ชื่อรุ่น]\n"
            "• ขอ3D Model : [ชื่อรุ่น]"
        )

    return jsonify({"ok": True}), 200


# =========================
# LINE Download Proxy
# GET /line/download?file=xxx.STEP&name=yyy.STEP
# =========================
@app.get("/line/download")
def line_download():
    from urllib.parse import unquote
    file_param = unquote(request.args.get("file", ""))
    name_param = unquote(request.args.get("name", "")) or file_param
    if not file_param:
        return jsonify({"error": "missing file param"}), 400
    file_bytes = _fetch_r2(file_param)
    if file_bytes is None:
        return jsonify({"error": f"ไม่พบไฟล์ {file_param}"}), 404
    import io
    from flask import send_file as _sf
    return _sf(
        io.BytesIO(file_bytes),
        as_attachment=True,
        download_name=name_param,
        mimetype="application/octet-stream",
    )


# =========================
# Health
# =========================

@app.get("/health")
def health():
    return jsonify(
        {
            "ok": True,
            "price_loaded": bool(PRICE_MAP),
            "price_count": len(PRICE_MAP),
            "price_file": PRICE_FILE,
            "template_file": TEMPLATE_FILE,
            "output_dir": str(OUTPUT_DIR),
        }
    )

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
